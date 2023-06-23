import re
from pdfminer.high_level import extract_text
from openpyxl import Workbook

# Відкрийте PDF-файл та отримайте його текст
text = extract_text('makinggames.pdf')

# Використовуйте регулярні вирази для знаходження електронних адрес
email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
emails = re.findall(email_pattern, text)

# Створіть нову Excel-таблицю
wb = Workbook()
ws = wb.active

# Додайте заголовки стовпців
ws['A1'] = 'Назва предприятия'
ws['B1'] = 'Електронна адреса'

# Запишіть дані у таблицю
for i, email in enumerate(emails, start=2):
    # Замість 'назва_предприятия' використовуйте код, який отримує назву предприятия
    ws[f'A{i}'] = 'SEXS'
    ws[f'B{i}'] = email

# Збережіть Excel-файл
wb.save('результат.xlsx')

